import subprocess
import os
import platform

os.system("pip install -r requirements.txt")

from openpyxl import Workbook, load_workbook

from certificate import *
from docx import Document
import csv
from docx2pdf import convert


mailerpath = "Data/Mail.xlsm"
htmltemplatepath = "Data/mailtemplate.html"

# create output folder if not exist
try:
    os.makedirs("Output/Doc")
    os.makedirs("Output/PDF")
except OSError:
    pass


def get_participants(f):
    data = [] # create empty list
    with open(f, mode="r", encoding='utf-8') as file:
        csv_reader = csv.DictReader(file)
        for row in csv_reader:
            data.append(row) # append all results
    return data

def updatemailer(row, workbook, sheet, email, filepath, sub, body, status, cc=""):
    sheet.cell(row=row, column=1).value = email
    sheet.cell(row=row, column=2).value = cc
    sheet.cell(row=row, column=3).value = sub
    sheet.cell(row=row, column=4).value = body
    sheet.cell(row=row, column=5).value = filepath
    sheet.cell(row=row, column=6).value = status
    workbook.save(filename = mailerpath)

def getworkbook(filename):
    wb = load_workbook(filename=filename, read_only=False, keep_vba=True)
    sheet = wb.active
    return wb, sheet

def gethtmltemplate(htmltemplatepath=htmltemplatepath):
    return open(htmltemplatepath, "r").read()

def getmail(name, event, lead, facilitator, gdsc):
    sub = f"[{event}] Certificate of Completion"

    html = gethtmltemplate(htmltemplatepath)

    short = "".join([i[0] for i in event.split(" ")])

    body = html.format(name=name, event=event, short=short, lead=lead, facilitator=facilitator, gdsc=gdsc)

    return sub, body

def convert_to_pdf(input_path, output_path):
    cmd = [
        "unoconv", 
        "-f", "pdf", 
        "-o", output_path, 
        input_path
    ]
    process = subprocess.Popen(cmd, stderr=subprocess.PIPE)
    output, error = process.communicate()
    return output, error

def create_docx_files(filename, list_participate, incomp=0, offset=2):

    wb, sheet = getworkbook(mailerpath)

    gdsc = input("Enter your college name: ")
    lead = input("Enter GDSC Lead Name: ")
    facilitator = input("Enter GCCP Facilitator Name: ")
    event = input("Enter the event name: ")

    for index, participate in enumerate(list_participate):
        # use original file everytime
        if participate["Total Completions of both Pathways"] == "No":
            filename = "Data/Event Certificate Template Participation.docx" # Certificate Template for participants
        elif participate["Total Completions of both Pathways"] == "Yes":
            filename = "Data/Event Certificate Template.docx" # Certificate Template for completions
        
        doc = Document(filename)

        name = participate["Student Name"]
        email = participate["Student Email"]

        replace_participant_name(doc, name)
        replace_gdsc_name(doc, gdsc)
        replace_lead_name(doc, lead)
        replace_facilitator_name(doc, facilitator)
        replace_event(doc, event)

        doc.save('Output/Doc/{}.docx'.format(name))

        # ! if your program working slowly, comment this two line and open other 2 line.
        print("Output/{}.pdf Creating".format(name))
        if platform.system() == 'Windows':
            convert('Output/Doc/{}.docx'.format(name), 'Output/Pdf/{}.pdf'.format(name))
        else:
            convert_to_pdf('Output/Doc/{}.docx'.format(name), 'Output/Pdf/{}.pdf'.format(name))

        filepath = os.path.abspath('Output/Pdf/{}.pdf'.format(name))

        sub, body = getmail(name, event, lead, facilitator, gdsc)

        updatemailer(row=(offset+index-incomp), workbook=wb,  sheet=sheet, email=email, filepath=filepath, sub=sub, body=body, status="Send")

    
# get certificate temple path
certificate_file = "Data/Event Certificate Template.docx"
# get participants path
participate_file = "Data/"+("ParticipantList.csv" if (input("Test Mode (Y/N): ").lower())[0]=="n" else "temp.csv")

# get participants
list_participate = get_participants(participate_file);

# process data
create_docx_files(certificate_file, list_participate)
